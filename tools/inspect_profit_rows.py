from __future__ import annotations

import json
import os
import sys

import openpyxl


def main() -> None:
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=False, read_only=False)
    ws = wb["Profit"]
    rows = []
    for row in range(25, 55):
        rows.append(
            {
                "row": row,
                "hidden": ws.row_dimensions[row].hidden,
                "A": ws.cell(row, 1).value,
                "B": ws.cell(row, 2).value,
                "D": ws.cell(row, 4).value,
            }
        )
    print(
        json.dumps(
            {
                "path": path,
                "mtime": os.path.getmtime(path),
                "sheets": wb.sheetnames,
                "profit_rows": rows,
            },
            default=str,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
