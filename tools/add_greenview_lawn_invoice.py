from __future__ import annotations

import copy
import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


TEAMS = Path(
    r"C:\Users\wesbr\Buy Your Home\Buy Your Home - Property\26_Project Management - 908 Pond St 3.xlsm"
)
ROOM = Path(
    r"C:\Codex\Wiki Files\Project Rooms\Project Management Spreadsheet Rewrite\working\invoice-project-updates"
)
SOURCE_FILE = (
    r"C:\Users\wesbr\Buy Your Home\Buy Your Home - Property\26-BYH -908 Pond St"
    r"\Owning\26-05-26 - Greenview Works.pdf"
)


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.font:
        dst.font = copy.copy(src.font)
    if src.border:
        dst.border = copy.copy(src.border)


def copy_column_width(ws, src_col: int, dst_col: int) -> None:
    ws.column_dimensions[get_column_letter(dst_col)].width = ws.column_dimensions[
        get_column_letter(src_col)
    ].width


def main() -> None:
    ROOM.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ROOM / f"26_Project Management - 908 Pond St 3 before Greenview lawn invoice {stamp}.xlsm"
    work = ROOM / "26_Project Management - 908 Pond St 3 - Greenview lawn invoice update.xlsm"
    shutil.copy2(TEAMS, backup)
    shutil.copy2(TEAMS, work)

    wb = openpyxl.load_workbook(work, keep_vba=True)
    ws = wb["Carrying"]
    profit = wb["Profit"]

    # Add a Lawn section after Excavator Rental, following the same three-column pattern.
    for src_col, dst_col in [(31, 34), (32, 35), (33, 36)]:
        copy_column_width(ws, src_col, dst_col)
    for row in range(1, 26):
        for src_col, dst_col in [(31, 34), (32, 35), (33, 36)]:
            copy_cell_style(ws.cell(row, src_col), ws.cell(row, dst_col))

    ws["AH1"] = "Lawn"
    ws["AH3"] = "Date"
    ws["AI3"] = "payment"
    ws["AH5"] = datetime(2026, 5, 26)
    ws["AI5"] = 60
    ws["AJ5"] = "Greenview Works invoice 000373"
    ws["AJ25"] = "=SUM(AI5:AI23)"
    ws["AH5"].number_format = "m/d/yyyy"
    ws["AI5"].number_format = "$#,##0.00"
    ws["AJ25"].number_format = "$#,##0.00"

    # Add Lawn to the existing Profit Carrying Cost block without otherwise converting the sheet.
    profit.insert_rows(40)
    for col in range(1, 5):
        copy_cell_style(profit.cell(39, col), profit.cell(40, col))
    profit["A40"] = "Lawn"
    profit["B40"] = "=+Carrying!AJ25/Profit!$B$28"
    profit["D40"] = None
    profit["A41"] = "Carrying Cost"
    profit["B41"] = "=+B28*SUM(B31:B40)"
    profit["D41"] = "=-B41"

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(work)

    # Basic verification from the saved workbook.
    check = openpyxl.load_workbook(work, keep_vba=True, data_only=False)
    c = check["Carrying"]
    p = check["Profit"]
    result = {
        "backup": str(backup),
        "work": str(work),
        "teams": str(TEAMS),
        "lawn_header": c["AH1"].value,
        "lawn_date": c["AH5"].value.isoformat() if c["AH5"].value else None,
        "lawn_amount": c["AI5"].value,
        "lawn_total_formula": c["AJ25"].value,
        "profit_lawn_label": p["A40"].value,
        "profit_lawn_formula": p["B40"].value,
        "profit_carrying_cost_label": p["A41"].value,
        "profit_carrying_cost_formula": p["B41"].value,
    }
    check.close()

    shutil.copy2(work, TEAMS)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
