from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


ROOT = Path(r"C:\Codex\Wiki Files\Project Rooms\Marketplace")
SOURCE = ROOT / "outputs" / "2026-08-14-estate-sale-click-review" / "estate-sale-click-review-2026-08-14.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "2026-08-14-estate-sale-price-labels-numbered"
OUTPUT_PDF = OUTPUT_DIR / "estate-sale-price-labels-numbered-avery-18660-2026-08-14.pdf"
REGISTER = ROOT / "working" / "estate-sale-item-number-register-2026-08-15.md"


# Existing sale-scoped identifiers must never be renumbered.
EXISTING_IDS = {
    "Assortment of Vinyl Records": 39,
    "Cased Sewing Machine Model 2468 with Notions": 30,
    "Decorative Doll and Angel Lot": 1,
    "Detecto 350-lb Physician Balance-Beam Scale": 40,
    "Estate Sale Bushnell PowerView 12x50 Binoculars": 67,
    "Estate Sale Crystal & Glassware Collection": 62,
    "Estate Sale Decorative Figurine & Music Box Collection": 76,
    "Estate Sale Nurse & Medical Collectible Figurine Lot": 75,
    "Estate Sale Poppytrail by Metlox Floral Dinnerware": 68,
    "Estate Sale Tony Little Gazelle Edge Exercise Glider": 72,
    "First Alert Locking Fire/Water Document Chest with Keys": 46,
    "Folding Drying-Rack Pair": 6,
    "Folding Rollaway Bed with Mattress": 36,
    "Folding Walker with Front Wheels": 21,
    "Four-Tier Basket Shelf": 16,
    "Frigidaire Top-Freezer Refrigerator": 8,
    "Fruit Decor and Green Oil-Lamp-Style Piece": 25,
    "Fur-Look Cape or Coat": 33,
    "GE Countertop Microwave": 4,
    "Handbag and Wallet Lot": 29,
    "Handled Wicker Basket": 26,
    "Household Iron with Metal Rest": 18,
    "Large Wood-Framed Dresser Mirror with Side Shelves": 44,
    "Lasko Digital Ceramic Heater": 13,
    "LG and GoVideo DVD-Player Pair": 32,
    "Marble-Look Pedestal Side Table": 23,
    "Maytag Washer and Whirlpool Dryer Pair": 7,
    "Mixed Decorative Figurine Lot": 37,
    "Mixed Hand-Tool Lot": 31,
    "Mixed Hardcover Book Lot": 28,
    "Nova 4215RD Red Rollator with Seat": 19,
    "Oak-Finish Desk with Hutch": 35,
    "Oak-Finish Entertainment Center": 10,
    "Oak-Finish Microwave Cabinet or Stand": 5,
    "Ornate Carved Mantel Clock": 38,
    "RCA RP-9515 Stereo and Speaker Set": 22,
    "Round Wood Stool": 12,
    "Sony CCD-F35 Video8 Camcorder Kit": 20,
    "Stained-Glass-Style Hanging Pendant Light": 2,
    "Tall Media Storage Cabinet": 14,
    "TIANSE 21-Hole Comb Binding Machine": 45,
    "Turquoise Table Lamp with Pleated Shade": 27,
    "Vintage Gold-Tone Rose-Pattern Flatware Set with Case": 42,
    "Wall-Mounted Toilet-Paper Storage Cabinet": 41,
    "Wood and Metal Bed Frame": 11,
    "Wood Corner Cabinet or Hutch": 24,
    "Wood Cubby Display Shelf with Figurines": 34,
    "Wood Folding TV Tray Table": 9,
    "Wood Lift-Top File or Record Cabinet": 15,
    "Wood Pedestal Book or Bible Stand": 17,
    "Wood Rolling Kitchen Cart": 3,
}


def read_prior_ids() -> dict[str, int]:
    if not REGISTER.exists():
        return {}
    prior: dict[str, int] = {}
    pattern = re.compile(r"^\| ES-20260815-(\d{3}) \| (.*?) \|")
    for line in REGISTER.read_text(encoding="utf-8").splitlines():
        match = pattern.match(line)
        if match:
            prior[match.group(2)] = int(match.group(1))
    return prior


def load_items() -> list[dict[str, object]]:
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook["Listing Analysis"]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values[1]:
            continue
        rows.append(
            {
                "category": str(values[0]),
                "title": str(values[1]),
                "price": int(values[2]),
                "status": str(values[14]),
            }
        )

    assigned = dict(EXISTING_IDS)
    assigned.update(read_prior_ids())
    used = set(assigned.values())
    next_id = 77
    for row in sorted(rows, key=lambda item: str(item["title"]).casefold()):
        title = str(row["title"])
        if title not in assigned:
            while next_id in used:
                next_id += 1
            assigned[title] = next_id
            used.add(next_id)
            next_id += 1
        row["number"] = assigned[title]

    numbers = [int(row["number"]) for row in rows]
    if len(numbers) != len(set(numbers)):
        raise ValueError("Duplicate item numbers found in the current label set")
    return sorted(rows, key=lambda item: int(item["number"]))


def wrap_title(text: str, max_width: float, preferred_size: float = 8.5) -> tuple[list[str], float]:
    for size in (preferred_size, 8.0, 7.5, 7.0):
        words = text.split()
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if stringWidth(candidate, "Helvetica-Bold", size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)
        if len(lines) <= 2:
            return lines, size
    return lines[:2], 7.0


def write_register(items: list[dict[str, object]]) -> None:
    lines = [
        "# Rosebrooks Estate Sale Item Number Register - 2026-08-15",
        "",
        "This register controls the stable reference numbers printed on the Avery 18660 price labels. Existing sale-scoped item numbers were preserved. Listings that did not previously have an item number were assigned the next unused number beginning at `077`. Do not reuse or renumber these references.",
        "",
        f"- Source pricing workbook: `{SOURCE.relative_to(ROOT)}`",
        f"- Labels: `{OUTPUT_PDF.relative_to(ROOT)}`",
        f"- Label count: {len(items)}",
        "- Price basis: current Facebook listing price fields captured 2026-08-14, including the $500 washer/dryer correction.",
        "",
        "| Item id | Listing / item | Category | Listed price | Facebook status |",
        "| --- | --- | --- | ---: | --- |",
    ]
    for item in items:
        lines.append(
            f"| ES-20260815-{int(item['number']):03d} | {item['title']} | {item['category']} | ${int(item['price']):,} | {item['status']} |"
        )
    lines.append("")
    REGISTER.write_text("\n".join(lines), encoding="utf-8")


def draw_label(pdf: canvas.Canvas, item: dict[str, object], x: float, y: float, width: float, height: float) -> None:
    inset = 7
    number = int(item["number"])
    status = str(item["status"])
    pdf.setFillColor(HexColor("#17365D"))
    pdf.rect(x, y + height - 15, width, 15, fill=1, stroke=0)
    pdf.setFillColor(white)
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(x + inset, y + height - 10.5, f"ITEM #{number:03d}")
    pdf.setFont("Helvetica-Bold", 6.2)
    pdf.drawRightString(x + width - inset, y + height - 10.2, "ROSEBROOKS ESTATE SALE")

    pdf.setFillColor(HexColor("#555555"))
    pdf.setFont("Helvetica-Bold", 5.8)
    category = str(item["category"]).upper()
    if stringWidth(category, "Helvetica-Bold", 5.8) > width - 2 * inset:
        category = category[:44]
    pdf.drawString(x + inset, y + height - 23, category)

    title_lines, title_size = wrap_title(str(item["title"]), width - 2 * inset)
    pdf.setFillColor(black)
    pdf.setFont("Helvetica-Bold", title_size)
    title_y = y + height - 34
    for line in title_lines:
        pdf.drawString(x + inset, title_y, line)
        title_y -= title_size + 1.2

    price_text = f"${int(item['price']):,}"
    pdf.setFont("Helvetica-Bold", 18)
    pdf.setFillColor(HexColor("#111111"))
    pdf.drawRightString(x + width - inset, y + 7.5, price_text)

    if status != "Active":
        pdf.setFillColor(HexColor("#B00020"))
        pdf.setFont("Helvetica-Bold", 6.7)
        pdf.drawString(x + inset, y + 8.5, "CHECK STATUS")


def write_pdf(items: list[dict[str, object]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    page_width, page_height = letter
    label_width = 2.625 * 72
    label_height = 1.0 * 72
    left_margin = 0.1875 * 72
    top_margin = 0.5 * 72
    horizontal_gap = 0.125 * 72
    labels_per_page = 30

    pdf = canvas.Canvas(str(OUTPUT_PDF), pagesize=letter, pageCompression=1)
    pdf.setTitle("Rosebrooks Estate Sale Numbered Price Labels - Avery 18660")
    pdf.setAuthor("Buy Your Home - Marketplace")
    for index, item in enumerate(items):
        slot = index % labels_per_page
        if index and slot == 0:
            pdf.showPage()
        row = slot // 3
        col = slot % 3
        x = left_margin + col * (label_width + horizontal_gap)
        y = page_height - top_margin - (row + 1) * label_height
        draw_label(pdf, item, x, y, label_width, label_height)
    pdf.save()


def main() -> None:
    items = load_items()
    if len(items) != 77:
        raise ValueError(f"Expected 77 current listings, found {len(items)}")
    write_register(items)
    write_pdf(items)
    print(f"Created {OUTPUT_PDF}")
    print(f"Updated {REGISTER}")
    print(f"Labels: {len(items)}; pages: {(len(items) + 29) // 30}; max item: {max(int(i['number']) for i in items):03d}")


if __name__ == "__main__":
    main()
