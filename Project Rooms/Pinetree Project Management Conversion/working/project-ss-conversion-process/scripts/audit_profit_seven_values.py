from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Pinetree Project Management Conversion")
converted = base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm"
source = base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"

targets = [
    ("Total CMA", "B4"),
    ("Total Purchase Cost", "K47"),
    ("Total Rehab Expense", "D67"),
    ("Total Debt", "K44"),
    ("Monthly Carrying Cost", "K41"),
    ("Monthly Income", "I55"),
    ("Total Profit", "I58"),
]

source_candidates = {
    "Total CMA": ["B4"],
    "Total Purchase Cost": ["K27"],
    "Total Rehab Expense": ["D51", "B30"],
    "Total Debt": ["E29", "B29"],
    "Monthly Carrying Cost": ["B27", "B20"],
    "Monthly Income": ["H42"],
    "Total Profit": ["I43", "K43", "K45"],
}

def open_pair(path):
    keep_vba = path.suffix.lower() == ".xlsm"
    fwb = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=keep_vba)
    vwb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    return fwb, vwb

def cell_info(ws_f, ws_v, addr):
    fc = ws_f[addr]
    vc = ws_v[addr]
    left = ws_f.cell(fc.row, max(1, fc.column - 1)).value
    label_a = ws_f.cell(fc.row, 1).value
    return {
        "addr": addr,
        "row_label_a": label_a,
        "left_label": left,
        "formula": fc.value,
        "cached_value": vc.value,
    }

cfwb, cvwb = open_pair(converted)
sfwb, svwb = open_pair(source)
cprofit_f, cprofit_v = cfwb["Profit"], cvwb["Profit"]
sprofit_f, sprofit_v = sfwb["Profit"], svwb["Profit"]

print(f"CONVERTED={converted}")
print(f"SOURCE={source}")
print()

for name, addr in targets:
    info = cell_info(cprofit_f, cprofit_v, addr)
    print(f"## {name}")
    print(f"converted: {info}")
    for candidate in source_candidates[name]:
        print(f"source_candidate: {cell_info(sprofit_f, sprofit_v, candidate)}")
    print()

cfwb.close()
cvwb.close()
sfwb.close()
svwb.close()
