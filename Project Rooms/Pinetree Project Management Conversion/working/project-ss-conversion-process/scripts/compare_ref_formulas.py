from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Pinetree Project Management Conversion")
files = {
    "pond": base / "template-reference" / "26_Project Management - 908 Pond St 3 - template.xlsm",
    "empty_labels_safe": base / "template-reference" / "Project Management - empty template from Pond Excel-native labels-safe 20260530_163443.xlsm",
    "pinetree_labels_safe": base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm",
}

sample_cells = [
    ("Others", "V7"),
    ("Others", "W7"),
    ("Others", "X7"),
    ("Gnatt Chart", "I2"),
    ("Gnatt Chart", "AE1"),
    ("Gnatt Chart", "AS1"),
    ("Gnatt Chart", "AU1"),
    ("Gnatt Chart", "BU7"),
    ("Draws", "R148"),
    ("Profit", "K80"),
    ("Profit", "K82"),
]

for label, path in files.items():
    wb = openpyxl.load_workbook(path, read_only=True, keep_vba=True, data_only=False)
    ref_count = 0
    formula_count = 0
    by_sheet = {}
    for ws in wb.worksheets:
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if "#REF!" in value:
                        ref_count += 1
                        count += 1
        if count:
            by_sheet[ws.title] = count
    print(f"\n[{label}] {path}")
    print(f"formulas={formula_count} formulas_with_ref={ref_count}")
    print("ref_by_sheet=" + ", ".join(f"{k}:{v}" for k, v in sorted(by_sheet.items())))
    for sheet, addr in sample_cells:
        if sheet in wb.sheetnames:
            print(f"{sheet}!{addr}: {wb[sheet][addr].value}")
    wb.close()
