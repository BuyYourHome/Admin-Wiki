from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Pinetree Project Management Conversion")
source = base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"
template = base / "template-reference" / "Project Management - empty template from Pond Excel-native labels-safe 20260530_163443.xlsm"

def dump(path, sheet, start, end):
    keep_vba = path.suffix.lower() == ".xlsm"
    fw = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=keep_vba)
    vw = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    fs, vs = fw[sheet], vw[sheet]
    print(f"\n## {path.name} {sheet} rows {start}:{end}")
    for r in range(start, end + 1):
        parts = []
        for c in range(1, 13):
            f = fs.cell(r, c).value
            v = vs.cell(r, c).value
            if f is None and v is None:
                continue
            coord = fs.cell(r, c).coordinate
            if isinstance(f, str) and f.startswith("="):
                parts.append(f"{coord}={f} [{v}]")
            else:
                parts.append(f"{coord}={f}")
        if parts:
            print("; ".join(parts))
    fw.close()
    vw.close()

dump(source, "Profit", 53, 82)
dump(template, "Profit", 69, 98)
