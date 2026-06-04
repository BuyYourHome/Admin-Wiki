from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Pinetree Project Management Conversion")
source = base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"
target = base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm"

def open_pair(path):
    keep_vba = path.suffix.lower() == ".xlsm"
    return (
        openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=keep_vba),
        openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba),
    )

def dump(path, label, sheet, ranges, cols=12):
    fw, vw = open_pair(path)
    fs, vs = fw[sheet], vw[sheet]
    print(f"\n## {label} {sheet}: {path}")
    for start, end in ranges:
        print(f"\nRows {start}:{end}")
        for r in range(start, end + 1):
            parts = []
            for c in range(1, cols + 1):
                f = fs.cell(r, c).value
                v = vs.cell(r, c).value
                if f is None and v is None:
                    continue
                coord = fs.cell(r, c).coordinate
                if isinstance(f, str) and f.startswith("="):
                    parts.append(f"{coord}={f} [{v}]")
                else:
                    parts.append(f"{coord}={f}")
            if parts:
                print("; ".join(parts))
    fw.close()
    vw.close()

dump(source, "SOURCE", "Profit", [(1, 10), (12, 18), (20, 31), (35, 45), (48, 51)], 12)
dump(target, "TARGET", "Profit", [(1, 10), (14, 21), (27, 28), (41, 48), (54, 58), (64, 67)], 12)
dump(source, "SOURCE", "Gnatt Chart", [(1, 8)], 12)
dump(target, "TARGET", "Gnatt Chart", [(1, 8)], 12)
