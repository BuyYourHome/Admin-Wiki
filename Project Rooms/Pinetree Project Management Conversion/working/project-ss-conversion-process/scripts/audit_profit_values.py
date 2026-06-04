from pathlib import Path
import re
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Pinetree Project Management Conversion")
converted = base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm"
source = base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"

keywords = [
    "cma",
    "purchase cost",
    "rehab",
    "debt",
    "carry",
    "income",
    "profit",
]

def norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()

def display(value):
    if value is None:
        return ""
    return str(value)

def load_pair(path):
    return (
        openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.lower() == ".xlsm"),
        openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=path.suffix.lower() == ".xlsm"),
    )

def row_snapshot(ws_formula, ws_values, row, min_col=1, max_col=12):
    cells = []
    for col in range(min_col, max_col + 1):
        fcell = ws_formula.cell(row, col)
        vcell = ws_values.cell(row, col)
        formula = fcell.value
        value = vcell.value
        text = display(formula if formula is not None else value)
        if text != "":
            cells.append(f"{fcell.coordinate}={text}")
    return "; ".join(cells)

def find_rows(path):
    wb_formula, wb_values = load_pair(path)
    ws_formula = wb_formula["Profit"]
    ws_values = wb_values["Profit"]
    hits = []
    seen_rows = set()
    for row in ws_formula.iter_rows():
        for cell in row:
            text = norm(cell.value)
            if not text:
                continue
            if any(k in text for k in keywords):
                if cell.row not in seen_rows:
                    seen_rows.add(cell.row)
                    hits.append((cell.row, row_snapshot(ws_formula, ws_values, cell.row)))
    wb_formula.close()
    wb_values.close()
    return hits

for label, path in [("SOURCE", source), ("CONVERTED", converted)]:
    print(f"\n## {label}: {path}")
    for row, snap in find_rows(path):
        print(f"Row {row}: {snap}")
