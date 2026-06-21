import json
import os
from datetime import datetime, time

import openpyxl
from openpyxl.utils import get_column_letter


SOURCE = r"C:\Codex\Wiki Files\Project Rooms\Project Management Spreadsheet Redesign\sources\17_Project Management - 3413 Pinetree Ln - source.xlsx"
TEMPLATE = os.environ["MODE2_TEMPLATE"]
OUTPUT_JSON = os.environ["MODE2_PAYLOAD"]

SHEETS = [
    "MOG", "Contact", "Seller Advance", "Demo & Trash Haul", "Appliances",
    "Plumbing Fixtures", "Windows & Doors", "Cabinets", "Paint", "Flooring",
    "HVAC", "Electrical Fixtures", "Landscape", "Exterior",
]


def payload_value(value):
    if isinstance(value, datetime):
        return {"kind": "date", "value": (value - datetime(1899, 12, 30)).days}
    if isinstance(value, time):
        return {"kind": "value", "value": value.isoformat()}
    return {"kind": "value", "value": value}


def main():
    src_formula = openpyxl.load_workbook(SOURCE, data_only=False, read_only=True)
    src_values = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    tmpl_formula = openpyxl.load_workbook(TEMPLATE, data_only=False, read_only=True, keep_vba=True)
    assignments = []

    for sheet in SHEETS:
        if sheet not in src_formula.sheetnames or sheet not in tmpl_formula.sheetnames:
            continue
        sf = src_formula[sheet]
        sv = src_values[sheet]
        tf = tmpl_formula[sheet]
        max_row = min(sf.max_row, tf.max_row)
        max_col = min(sf.max_column, tf.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                src_cell = sf.cell(row, col)
                tgt_cell = tf.cell(row, col)
                if isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                    continue
                if isinstance(tgt_cell.value, str) and tgt_cell.value.startswith("="):
                    continue
                value = sv.cell(row, col).value
                if value is None:
                    continue
                assignments.append({
                    "sheet": sheet,
                    "addr": f"{get_column_letter(col)}{row}",
                    **payload_value(value),
                })

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump({"assignments": assignments}, f)
    print(json.dumps({"assignment_count": len(assignments), "output": OUTPUT_JSON}))


if __name__ == "__main__":
    main()
