from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Project Management Spreadsheet Redesign")
files = [
    ("source", base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"),
    ("converted", base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm"),
]
ranges = {
    "source": [(20, 31), (35, 46), (48, 51)],
    "converted": [(21, 28), (41, 48), (54, 58), (64, 67)],
}

def open_pair(path):
    keep_vba = path.suffix.lower() == ".xlsm"
    return (
        openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=keep_vba),
        openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba),
    )

for label, path in files:
    fwb, vwb = open_pair(path)
    fws, vws = fwb["Profit"], vwb["Profit"]
    print(f"\n## {label}: {path}")
    for start, end in ranges[label]:
        print(f"\nRows {start}:{end}")
        for row in range(start, end + 1):
            pieces = []
            for col in range(1, 12):
                fc = fws.cell(row, col)
                vc = vws.cell(row, col)
                if fc.value is not None or vc.value is not None:
                    value = fc.value
                    cached = vc.value
                    if isinstance(value, str) and value.startswith("="):
                        pieces.append(f"{fc.coordinate} {value} [{cached}]")
                    else:
                        pieces.append(f"{fc.coordinate} {value}")
            print("; ".join(pieces))
    fwb.close()
    vwb.close()
