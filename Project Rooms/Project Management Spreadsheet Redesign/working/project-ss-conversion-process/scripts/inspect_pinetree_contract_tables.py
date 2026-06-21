from pathlib import Path
import openpyxl

base = Path(r"C:\Codex\Wiki Files\Project Rooms\Project Management Spreadsheet Redesign")
target = base / "Need Verification" / "17_Project Management - 3413 Pinetree Ln - mode2 labels-safe 20260530_163443.xlsm"
source = base / "sources" / "17_Project Management - 3413 Pinetree Ln - source.xlsx"

def show(path, sheet, min_row, max_row, max_col=12):
    keep_vba = path.suffix.lower() == ".xlsm"
    fw = openpyxl.load_workbook(path, read_only=True, data_only=False, keep_vba=keep_vba)
    vw = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=keep_vba)
    fs, vs = fw[sheet], vw[sheet]
    print(f"\n## {path.name} :: {sheet} rows {min_row}:{max_row}")
    for r in range(min_row, max_row + 1):
        parts = []
        for c in range(1, max_col + 1):
            f = fs.cell(r, c).value
            v = vs.cell(r, c).value
            if f is None and v is None:
                continue
            coord = fs.cell(r, c).coordinate
            if isinstance(f, str) and f.startswith("="):
                parts.append(f"{coord}={f} [{v}]")
            else:
                parts.append(f"{coord}={f}")
        if parts:
            print("; ".join(parts))
    fw.close()
    vw.close()

for sh in ["Contract", "ProposalUpdate", "Demo & Trash Haul", "Appliances", "Plumbing Fixtures", "Windows & Doors", "Cabinets", "Paint", "Flooring", "HVAC", "Electrical Fixtures", "Landscape", "Exterior"]:
    try:
        show(target, sh, 1, 25, 12)
    except Exception as e:
        print(f"skip target {sh}: {e}")

for sh in ["ProposalUpdate", "Demo & Trash Haul", "Appliances", "Plumbing Fixtures", "Windows & Doors", "Cabinets", "Paint", "Flooring", "HVAC", "Electrical Fixtures", "Landscape", "Exterior"]:
    try:
        show(source, sh, 1, 25, 12)
    except Exception as e:
        print(f"skip source {sh}: {e}")
